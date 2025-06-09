import os
import numpy as np
import matplotlib.pyplot as plt
import data_reader as dr
import data_processor as dp

from openpyxl import Workbook
from scipy.signal import savgol_filter


# 创建拟合结果记录表
def create_output_sheet(wb, sample_name, sam_index, columns):
    ws = wb.create_sheet(sample_name)
    ws.freeze_panes = 'A3'
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(2, col_idx).value = col_name
    ws.cell(1, 1).hyperlink = f"#'Sample list'!A{sam_index}"
    return ws

# 回写拟合结果记录表
def pushpack_peak(worksheet, cont, col_index, start_row=3):
    j = start_row
    for cell in cont:
        worksheet.cell(j, col_index).value = float(cell) if cell is not None else None
        j += 1
    return worksheet

if __name__ == "__main__":
    # 选择文件类型
    file_type = input("请选择文件类型 (1: Philips.rd, 2: Panalytical.xrdml, 3: Rigaku.raw): ").strip()
    calc_type = input("请选择计算类型 (1: D002, 2: Si_FWHM, 3: OI): ").strip()
    if calc_type == "2":
        smooth_y = input("是否平滑数据 (1: Yes, 2: No): ").strip()
    peak_output = input("是否输出拟合结果 (1: Yes, 2: No): ").strip()
    if peak_output == "1":
        trim_filename = input("是否裁切文件名 (1: Yes, 2: No): ").strip()
    else:
        trim_filename = "0"
    reader = dr.get_reader(file_type)

    # 抓取文件清单
    file_list = []
    for root, _, files in os.walk('.'):
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if (file_type == "1" and ext == ".rd") or (file_type == "2" and ext == ".xrdml") or (file_type == "3" and ext == ".raw"):
                file_list.append(os.path.join(root, f))

    # 初始化计算结果汇总表
    wb = Workbook()
    ws_res = wb.active
    ws_res.title = 'Sample list'
    ws_res.freeze_panes = 'A2'
    if calc_type == "1":
        ws_res.cell(1, 1).value = 'Sample Name'
        ws_res.cell(1, 2).value = 'D002 (Å)'
        ws_res.cell(1, 3).value = 'G%'
        ws_res.cell(1, 4).value = 'Graphite [002] Peak (deg)'
        ws_res.cell(1, 5).value = 'Graphite [002] FWHM (deg)'
        ws_res.cell(1, 6).value = 'Silicon [111] Peak (deg)'
        ws_res.cell(1, 7).value = 'Silicon [111] FWHM (deg)'
        ws_res.cell(1, 8).value = 'Graphite [002] FWHM NET.(deg)'
        ws_res.cell(1, 9).value = 'Graphite [002] Lc NET.(Å)'
        ws_res.cell(1, 10).value = 'Graphite [002] FWHM JIS(deg)'
        ws_res.cell(1, 11).value = 'Graphite [002] Lc JIS(Å)'
    elif calc_type == "2":
        ws_res.cell(1, 1).value = 'Sample Name'
        ws_res.cell(1, 2).value = 'Silicon [111] Peak (deg)'
        ws_res.cell(1, 3).value = 'Silicon [111] FWHM (deg)'
    elif calc_type == "3":
        ws_res.cell(1, 1).value = 'Sample Name'
        ws_res.cell(1, 2).value = 'Graphite OI value (-)'
        ws_res.cell(1, 3).value = 'Graphite [004] Peak (deg)'
        ws_res.cell(1, 4).value = 'Graphite [004] Intensity (counts)'
        ws_res.cell(1, 5).value = 'Graphite [004] FWHM (deg)'
        ws_res.cell(1, 6).value = 'Graphite [110] Peak (deg)'
        ws_res.cell(1, 7).value = 'Graphite [110] Intensity (counts)'
        ws_res.cell(1, 8).value = 'Graphite [110] FWHM (deg)'

    # 初始化辅助变量
    sam_list = []
    sam_index = 1

    # 数据处理模块
    for file_path in file_list:
        sample_name = os.path.splitext(os.path.basename(file_path))[0]
        sam_list.append(sample_name)
        sam_index += 1
        try:
            scan_x, scan_y = reader.read_data(file_path)
            if (scan_x is not None)&(calc_type == "1"):
                # 计算石墨 D002 相关
                corrected_x, corrected_y = dp.correct_ka2(scan_x, scan_y)
                popt = dp.fit_data_d002(corrected_x, corrected_y)
                if popt is not None:
                    fitted_curve, background, graphite_peak, silicon_peak, fwhm_gn = dp.fit_peak_d002(corrected_x, popt)  # 计算拟合结果曲线
                    ws_res.cell(sam_index, 1).value = sample_name  # 样品名
                    ws_res.cell(sam_index, 2).value = d_002 = 1.54056/(2*np.sin(np.radians((28.443 - popt[6] + popt[1])/2)))  # 计算石墨 D002 层间距
                    ws_res.cell(sam_index, 3).value = 100 * (3.440 - d_002)/(3.440 - 3.354)  # 计算石墨化度
                    ws_res.cell(sam_index, 4).value = popt[1]  # 计算石墨 [002] 峰位
                    ws_res.cell(sam_index, 5).value = fwhm_g = dp.calculate_fwhm_spv(popt[2], popt[3], popt[4])  # 计算石墨 [002] 峰半峰宽
                    ws_res.cell(sam_index, 6).value = popt[6]  # 计算硅 [111] 峰位
                    ws_res.cell(sam_index, 7).value = fwhm_si = dp.calculate_fwhm_spv(popt[7], popt[8], popt[9])  # 计算硅 [111] 峰半峰宽
                    ws_res.cell(sam_index, 8).value = fwhm_gn  # 计算石墨半峰宽 via deconvolution
                    ws_res.cell(sam_index, 9).value = 0.89 * 1.54056 / (np.radians(fwhm_gn) * np.cos(np.radians((28.443 - popt[6] + popt[1])/2)))  # 计算石墨Lc值 via deconvolution
                    ws_res.cell(sam_index, 10).value = fwhm_jis = dp.calculate_fwhm_jis(fwhm_g, fwhm_si)  # 计算半峰宽 via JISR7651:2007
                    ws_res.cell(sam_index, 11).value = 0.89 * 1.54056 / (np.radians(fwhm_jis) * np.cos(np.radians((28.443 - popt[6] + popt[1])/2)))  # 计算Lc值 via JISR7651:2007
                    if peak_output == "1":
                        # 初始化拟合结果记录表
                        if trim_filename == "1":
                            try:
                                sample_name = sample_name.split()[-3]
                            except:
                                pass
                        ws_raw = wb.create_sheet(sample_name)
                        ws_raw.freeze_panes = 'A3'
                        ws_raw.cell(2, 1).value = '2-Theta (deg)'
                        ws_raw.cell(2, 2).value = 'Intensity (A.U.)'
                        ws_raw.cell(2, 3).value = 'Corrected Intensity'
                        ws_raw.cell(2, 4).value = 'Fitted Curve'
                        ws_raw.cell(2, 5).value = 'Fitted Background'
                        ws_raw.cell(2, 6).value = 'Graphite [002] Net. Peak'
                        ws_raw.cell(2, 7).value = 'Silicon [111s] Net. Peak'
                        ws_raw.cell(2, 8).value = 'Graphite [002] Peak'
                        ws_raw.cell(2, 9).value = 'Silicon [111] Peak'
                        ws_raw.cell(1, 1).value = 'Back'
                        ws_raw.cell(1, 1).hyperlink = "#'Sample list'!A%s" % str(sam_index)
                        ws_res.cell(sam_index, 1).hyperlink = "#'%s'!A1" %(sample_name)
                        # 回写拟合结果曲线
                        ws_raw = pushpack_peak(ws_raw, corrected_x, 1)
                        ws_raw = pushpack_peak(ws_raw, scan_y, 2)
                        ws_raw = pushpack_peak(ws_raw, corrected_y, 3)
                        ws_raw = pushpack_peak(ws_raw, fitted_curve, 4)
                        ws_raw = pushpack_peak(ws_raw, background, 5)
                        ws_raw = pushpack_peak(ws_raw, graphite_peak, 6)
                        ws_raw = pushpack_peak(ws_raw, silicon_peak, 7)
                        ws_raw = pushpack_peak(ws_raw, graphite_peak + background, 8)
                        ws_raw = pushpack_peak(ws_raw, silicon_peak + background, 9)
                        # 创建绘图
                        plt.figure(figsize=(10, 6))  # 设置图表大小
                        # plt.plot(corrected_x, scan_y, label='Raw Intensity', linewidth=0.5, color='gray', linestyle='--')  # 原始数据
                        plt.plot(corrected_x, corrected_y, label='Corrected Intensity', linewidth=0.5, color='blue')  # 校正后数据
                        plt.plot(corrected_x, fitted_curve, label='Fitted Curve', linewidth=0.5, color='red')  # 拟合曲线
                        plt.plot(corrected_x, background, label='Background', linewidth=0.5, color='green', linestyle='--')  # 背景
                        plt.plot(corrected_x, graphite_peak + background, label='Graphite [002] Peak', linewidth=0.5, color='orange')  # 石墨峰
                        plt.plot(corrected_x, silicon_peak + background, label='Silicon [111] Peak', linewidth=0.5, color='purple')  # 硅峰

                        # 美化图表
                        plt.title(sample_name, fontsize=14, fontweight='bold')  # 标题
                        plt.xlabel('2Theta (°)', fontsize=12)  # X轴标签
                        plt.xlim(25, 29)
                        plt.ylabel('Intensity (Counts)', fontsize=12)  # Y轴标签
                        plt.grid(True, linestyle='--', linewidth=0.5, alpha=0.7)  # 网格线
                        plt.legend(loc='upper left', fontsize=10)  # 图例
                        plt.tight_layout()  # 自动调整布局

                        # 保存图表
                        plt.savefig(f'{sample_name}_plot.png', dpi=300)  # 保存为PNG文件
                        plt.close()  # 关闭图表，释放内存


                else: 
                    print(f"Failed to fit peaks for sample {sample_name}: 拟合失败，未返回参数")
                    
            elif (scan_x is not None)&(calc_type == "2"):
                # 计算纳米硅相关
                corrected_x, corrected_y = dp.correct_ka2(scan_x, scan_y)
                smoothed_y = savgol_filter(corrected_y, 25, 3)
                if smooth_y == "1":
                    popt = dp.fit_data_sifwhm(corrected_x, smoothed_y)
                else:
                    popt = dp.fit_data_sifwhm(corrected_x, corrected_y)
                if popt is not None:
                    ws_res.cell(sam_index, 1).value = sample_name  # 样品名
                    ws_res.cell(sam_index, 2).value = popt[1]  # 计算硅 [111] 峰位
                    ws_res.cell(sam_index, 3).value = dp.calculate_fwhm_spv(popt[2], popt[3], popt[4])  # 计算硅 [111] 峰半峰宽
                    if peak_output == "1":
                        # 初始化拟合结果记录表
                        if trim_filename == "1":
                            try:
                                sample_name = sample_name.split()[-3]
                            except:
                                pass
                        ws_raw = wb.create_sheet(sample_name)
                        ws_raw.freeze_panes = 'A3'
                        ws_raw.cell(2, 1).value = '2-Theta (deg)'
                        ws_raw.cell(2, 2).value = 'Intensity (A.U.)'
                        ws_raw.cell(2, 3).value = 'Corrected Intensity'
                        ws_raw.cell(2, 4).value = 'Smoothed Intensity'
                        ws_raw.cell(2, 5).value = 'Fitted Curve'
                        ws_raw.cell(2, 6).value = 'Fitted Background'
                        ws_raw.cell(2, 7).value = 'Silicon [111] Peak'
                        ws_raw.cell(1, 1).value = 'Back'
                        ws_raw.cell(1, 1).hyperlink = "#'Sample list'!A%s" % str(sam_index)
                        ws_res.cell(sam_index, 1).hyperlink = "#'%s'!A1" %(sample_name)
                        # 计算拟合结果曲线
                        fitted_curve, background, silicon_peak = dp.fit_peak_sifwhm(corrected_x, popt)
                        # 回写拟合结果曲线
                        ws_raw = pushpack_peak(ws_raw, corrected_x, 1)
                        ws_raw = pushpack_peak(ws_raw, scan_y, 2)
                        ws_raw = pushpack_peak(ws_raw, corrected_y, 3)
                        ws_raw = pushpack_peak(ws_raw, smoothed_y, 4)
                        ws_raw = pushpack_peak(ws_raw, fitted_curve, 5)
                        ws_raw = pushpack_peak(ws_raw, background, 6)
                        ws_raw = pushpack_peak(ws_raw, silicon_peak, 7)
                else:
                    print(f"Failed to fit peaks for sample {sample_name}: 拟合失败，未返回参数")
                    ws_res.cell(sam_index, 1).value = sample_name  # 样品名
                    if peak_output == "1":
                        # 初始化拟合结果记录表
                        if trim_filename == "1":
                            try:
                                sample_name = sample_name.split()[-3]
                            except:
                                pass
                        ws_raw = wb.create_sheet(sample_name)
                        ws_raw.freeze_panes = 'A3'
                        ws_raw.cell(2, 1).value = '2-Theta (deg)'
                        ws_raw.cell(2, 2).value = 'Intensity (A.U.)'
                        ws_raw.cell(2, 3).value = 'Corrected Intensity'
                        ws_raw.cell(2, 4).value = 'Smoothed Intensity'
                        ws_raw.cell(1, 1).value = 'Back'
                        ws_raw.cell(1, 1).hyperlink = "#'Sample list'!A%s" % str(sam_index)
                        ws_res.cell(sam_index, 1).hyperlink = "#'%s'!A1" %(sample_name)
                        # 回写拟合结果曲线
                        ws_raw = pushpack_peak(ws_raw, corrected_x, 1)
                        ws_raw = pushpack_peak(ws_raw, scan_y, 2)
                        ws_raw = pushpack_peak(ws_raw, corrected_y, 3)
                        ws_raw = pushpack_peak(ws_raw, smoothed_y, 4)

            elif (scan_x is not None)&(calc_type == "3"):
                # 计算OI值
                corrected_x, corrected_y = dp.correct_ka2(scan_x, scan_y)
                popt = dp.fit_data_oi(corrected_x, corrected_y)
                if popt is not None:
                    ws_res.cell(sam_index, 1).value = sample_name  # 样品名
                    ws_res.cell(sam_index, 2).value = popt[0]/popt[20]  # 计算OI值
                    ws_res.cell(sam_index, 3).value = popt[1]  # 计算石墨 [004] 峰位
                    ws_res.cell(sam_index, 4).value = popt[0]  # 计算石墨 [004] 峰高
                    ws_res.cell(sam_index, 5).value = dp.calculate_fwhm_spv(popt[2], popt[3], popt[4])  # 计算石墨 [004] 峰半峰宽
                    ws_res.cell(sam_index, 6).value = popt[21]  # 计算石墨 [110] 峰位
                    ws_res.cell(sam_index, 7).value = popt[20]  # 计算石墨 [110] 峰高
                    ws_res.cell(sam_index, 8).value = dp.calculate_fwhm_spv(popt[22], popt[23], popt[24])  # 计算石墨 [110] 峰半峰宽
                    if peak_output == "1":
                        # 初始化拟合结果记录表
                        if trim_filename == "1":
                            try:
                                sample_name = sample_name.split()[-3]
                            except:
                                pass
                        ws_raw = wb.create_sheet(sample_name)
                        ws_raw.freeze_panes = 'A3'
                        ws_raw.cell(2, 1).value = '2-Theta (deg)'
                        ws_raw.cell(2, 2).value = 'Intensity (A.U.)'
                        ws_raw.cell(2, 3).value = 'Corrected Intensity'
                        ws_raw.cell(2, 4).value = 'Fitted Curve'
                        ws_raw.cell(2, 5).value = 'Fitted Background'
                        ws_raw.cell(2, 6).value = 'Graphite [004] Peak'
                        ws_raw.cell(2, 7).value = 'Silicon [311] Peak'
                        ws_raw.cell(2, 8).value = 'Silicon [400] Peak'
                        ws_raw.cell(2, 9).value = 'Silicon [331] Peak'
                        ws_raw.cell(2, 10).value = 'Graphite [110] Peak'
                        ws_raw.cell(1, 1).value = 'Back'
                        ws_raw.cell(1, 1).hyperlink = "#'Sample list'!A%s" % str(sam_index)
                        ws_res.cell(sam_index, 1).hyperlink = "#'%s'!A1" %(sample_name)
                        # 计算拟合结果曲线
                        fitted_curve, background, g004_peak, si311_peak, si400_peak, si331_peak, g110_peak = dp.fit_peak_oi(corrected_x, popt)
                        # 回写拟合结果曲线
                        ws_raw = pushpack_peak(ws_raw, corrected_x, 1)
                        ws_raw = pushpack_peak(ws_raw, scan_y, 2)
                        ws_raw = pushpack_peak(ws_raw, corrected_y, 3)
                        ws_raw = pushpack_peak(ws_raw, fitted_curve, 4)
                        ws_raw = pushpack_peak(ws_raw, background, 5)
                        ws_raw = pushpack_peak(ws_raw, g004_peak, 6)
                        ws_raw = pushpack_peak(ws_raw, si311_peak, 7)
                        ws_raw = pushpack_peak(ws_raw, si400_peak, 8)
                        ws_raw = pushpack_peak(ws_raw, si331_peak, 9)
                        ws_raw = pushpack_peak(ws_raw, g110_peak, 10)
                else: print(f"Failed to fit peaks for sample {sample_name}: 拟合失败，未返回参数")
            else:
                print(f"Failed to fit peaks for sample {sample_name}: 拟合失败，文件读取错误")
        except RuntimeError as e:
            print(f"Failed to fit peaks for sample {sample_name}: {e}")

    wb.save('xrd_processed.xlsx')
    print('处理完成，结果已写入')
